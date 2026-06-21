"""
Importa o histórico da planilha antiga para o banco do sistema.

Uso:
    python -m scripts.import_planilha /caminho/Tiro_Guerra_AAAAMMDD.xlsx

Cada aba de mês (ex.: "26 MAI") vira um mês no sistema, com suas etapas,
inscrições e resultados. Os tempos históricos já incluíam as penalidades,
então são importados como o "tempo final" de cada corrida — o que mantém
as pontuações idênticas às da planilha.
"""
import re
import sys
from datetime import date

from openpyxl import load_workbook

from app import db

# Aba -> (ano, mês). Ajuste aqui se mudar a nomenclatura das abas.
SHEET_MONTHS = {
    "26 FEV": (2026, 2),
    "26 MAR": (2026, 3),
    "26 ABR": (2026, 4),
    "26 MAI": (2026, 5),
    "26 JUN": (2026, 6),
}
# Ordem cronológica de processamento (define a categoria "atual" pelo mês mais recente)
ORDER = ["26 FEV", "26 MAR", "26 ABR", "26 MAI", "26 JUN"]

# Colunas de "Tempo" de cada etapa (1-indexed): D,F,H,J,L
TIME_COLS = [4, 6, 8, 10, 12]


def norm(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip()).upper()


def map_category(conn, label: str):
    """Retorna (modality, category_id) a partir do rótulo do bloco na planilha."""
    u = norm(label)
    if "CARABINA" in u:
        modality = "carabina"
        cat_name = "Geral"
    elif "VETERANO" in u:
        modality, cat_name = "pistola", "Veterano de Guerra"
    elif "GUERREIRO" in u:
        modality, cat_name = "pistola", "Guerreiro"
    elif "COMBATENTE" in u:
        modality, cat_name = "pistola", "Combatente"
    else:
        return None, None
    row = conn.execute(
        "SELECT id FROM categories WHERE modality=? AND name=?",
        (modality, cat_name)).fetchone()
    return modality, (row["id"] if row else None)


def main(xlsx_path: str):
    db.init_db()
    conn = db.get_conn()
    wb = load_workbook(xlsx_path, data_only=True)

    shooter_cache = {}  # norm(name) -> shooter_id

    def get_or_create_shooter(name):
        key = norm(name)
        if key in shooter_cache:
            return shooter_cache[key]
        existing = conn.execute(
            "SELECT id FROM shooters WHERE UPPER(name)=?", (key,)).fetchone()
        if existing:
            shooter_cache[key] = existing["id"]
            return existing["id"]
        sid = db.create_shooter(conn, str(name).strip())
        shooter_cache[key] = sid
        return sid

    total_runs = 0
    for sheet in ORDER:
        if sheet not in wb.sheetnames:
            continue
        year, month = SHEET_MONTHS[sheet]
        ws = wb[sheet]

        # mês
        conn.execute(
            "INSERT OR IGNORE INTO months(year,month,status,created_at) "
            "VALUES (?,?,?,?)", (year, month, "aberto", db.now()))
        m = conn.execute("SELECT * FROM months WHERE year=? AND month=?",
                         (year, month)).fetchone()
        month_id = m["id"]

        # datas das etapas (linha 2)
        stage_ids = {}
        for idx, col in enumerate(TIME_COLS, start=1):
            d = ws.cell(row=2, column=col).value
            if d is None:
                continue
            the_date = d.date().isoformat() if hasattr(d, "date") else str(d)
            conn.execute(
                "INSERT OR IGNORE INTO stages(month_id,number,date,status,created_at) "
                "VALUES (?,?,?,?,?)", (month_id, idx, the_date, "fechada", db.now()))
            srow = conn.execute(
                "SELECT id FROM stages WHERE month_id=? AND number=?",
                (month_id, idx)).fetchone()
            stage_ids[idx] = srow["id"]

        # percorre as linhas de dados
        current_modality, current_cat = None, None
        for r in range(5, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value      # categoria (início do bloco)
            name = ws.cell(row=r, column=3).value   # atirador
            if a and ("PARTICIPANTES" in norm(a)):
                break
            if a:
                mod, cat = map_category(conn, a)
                if cat:
                    current_modality, current_cat = mod, cat
            if not name or current_cat is None:
                continue
            if "PARTICIPANTES" in norm(name):
                continue
            sid = get_or_create_shooter(name)
            # categoria do mês (snapshot) e categoria atual
            db.set_shooter_category(conn, sid, current_modality, current_cat)
            conn.execute(
                "INSERT OR REPLACE INTO month_category"
                "(month_id,shooter_id,modality,category_id) VALUES (?,?,?,?)",
                (month_id, sid, current_modality, current_cat))

            runs_here = 0
            for idx, col in enumerate(TIME_COLS, start=1):
                if idx not in stage_ids:
                    continue
                t = ws.cell(row=r, column=col).value
                if t is None or not isinstance(t, (int, float)) or t <= 0:
                    continue
                # inscrição
                enr = db.get_enrollment(conn, stage_ids[idx], sid, current_modality)
                if not enr:
                    enr = db.enroll(conn, stage_ids[idx], sid, current_modality,
                                    current_cat, 1, 0)
                # corrida (tempo já é o tempo final histórico)
                db.add_run(conn, enr["id"], float(t), 0, 0, 0, False, 0)
                runs_here += 1
                total_runs += 1

    conn.commit()

    # status dos meses: o mês corrente fica aberto, os demais fechados
    today = date.today()
    conn.execute("UPDATE months SET status='fechado'")
    conn.execute("UPDATE months SET status='aberto' WHERE year=? AND month=?",
                 (today.year, today.month))
    # garante o mês corrente existindo e aberto
    db.ensure_current_month(conn)
    conn.execute("UPDATE months SET status='aberto' WHERE year=? AND month=?",
                 (today.year, today.month))
    conn.commit()

    n_shooters = conn.execute("SELECT COUNT(*) c FROM shooters").fetchone()["c"]
    n_months = conn.execute("SELECT COUNT(*) c FROM months").fetchone()["c"]
    conn.close()
    print(f"Importação concluída: {n_shooters} atiradores, {n_months} meses, "
          f"{total_runs} resultados.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python -m scripts.import_planilha caminho/arquivo.xlsx")
        sys.exit(1)
    main(sys.argv[1])
